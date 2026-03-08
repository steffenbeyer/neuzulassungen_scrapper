"""
Abstrakte Basisklasse fuer alle KBA Excel-Parser.
"""
import logging
from abc import ABC, abstractmethod
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class BaseParser(ABC):
    """Basisklasse fuer KBA Excel-Parser."""

    # Wird von Unterklassen ueberschrieben
    QUELLE_KUERZEL = None

    def __init__(self):
        self.workbook = None
        self.filepath = None

    def load(self, filepath):
        """Laedt eine Excel-Datei."""
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            raise FileNotFoundError(f"Datei nicht gefunden: {filepath}")

        logger.info(f"Lade Excel: {self.filepath.name}")
        self.workbook = load_workbook(str(self.filepath), read_only=True, data_only=True)
        return self

    def get_sheet_names(self):
        """Gibt alle Sheet-Namen zurueck."""
        if not self.workbook:
            return []
        return self.workbook.sheetnames

    def get_sheet(self, name=None, index=0):
        """Gibt ein Worksheet zurueck (nach Name oder Index)."""
        if not self.workbook:
            return None
        if name:
            return self.workbook[name] if name in self.workbook.sheetnames else None
        return self.workbook.worksheets[index] if index < len(self.workbook.worksheets) else None

    def extract_year_month_from_filename(self):
        """Extrahiert Jahr und Monat aus dem Dateinamen (z.B. fz10_2024_01.xlsx)."""
        stem = self.filepath.stem  # z.B. "fz10_2024_01"
        parts = stem.split('_')
        try:
            if len(parts) >= 3:
                year = int(parts[1])
                month = int(parts[2])
                return year, month
            elif len(parts) == 2:
                year = int(parts[1])
                return year, None
        except (ValueError, IndexError):
            pass

        logger.warning(f"Kann Jahr/Monat nicht aus Dateiname extrahieren: {self.filepath.name}")
        return None, None

    @abstractmethod
    def parse(self, filepath):
        """
        Parst eine Excel-Datei und gibt strukturierte Daten zurueck.

        Args:
            filepath: Pfad zur Excel-Datei

        Returns:
            list: Liste von Dicts mit den geparsten Daten
        """
        pass

    def close(self):
        """Schliesst die geladene Excel-Datei."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
